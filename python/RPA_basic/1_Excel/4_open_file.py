from openpyxl import load_workbook

wb = load_workbook("cells.xlsx")  # create_file.xlsx  파일에서 wb 불러오기
ws = wb.active  # 활성화된 sheet

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ...
#     print()

# cell의 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ...
    print()

